import random
from os import path
import pandas as pd
from datetime import date
import openpyxl
from openpyxl.styles import PatternFill

# ------------------------------------------- GLOBALS ------------------------------------------- #

today = date.today().strftime("%d/%m/%y")

waiters = ['Lesly', 'Ismail', 'Alex', 'Hakan', 'Alexandra','Yolanda', 'Pragya', 'Maroon', 'Sivu', 'Kevin', 'Patrick']
sections = {
    'Table1': {'location': 'outside', 'capacity': 1, 'importance': True},
    'Table2-3': {'location': 'outside', 'capacity': 1, 'importance': True},
    'Table4-5': {'location': 'outside', 'capacity': 1, 'importance': True},
    'NewSection': {'location': 'outside', 'capacity': 3, 'importance': True},
    'Middle': {'location': 'inside', 'capacity': 2, 'importance': False},
    '13-14-15': {'location': 'inside', 'capacity': 1, 'importance': True},
    '16-17-18': {'location': 'inside', 'capacity': 1, 'importance': True},
    'SmokingSection': {'location': 'inside', 'capacity': 2, 'importance': False},
}
priority_waiters = []
unimportant_sections = []
important_sections = []
outside_sections = []
inside_sections = []
full_sections = []
important_section_capacity = 0
high_capacity_sections = []
unimportant_section_capacity = 0
section_assignments = dict.fromkeys(sections.keys(), None)
assigned_waiters = []
for section_key, section_values in sections.items():
    if section_values["location"] == "outside":
        outside_sections.append(section_key)
    else:
        inside_sections.append(section_key)
    if not section_values["importance"]:
        unimportant_sections.append(section_key)
        unimportant_section_capacity += section_values['capacity']
    else:
        important_section_capacity += section_values["capacity"]
        important_sections.append(section_key)
    if section_values['capacity'] > 1:
        high_capacity_sections.append(section_key)
try:
    waiter_history_df = pd.read_excel("./Data/RotationCopy.xlsx", index_col=0)
    waiter_history_dict = waiter_history_df.to_dict("list")
except:
    if not path.exists("./Data/Rotation.xlsx"):
        rotation = pd.DataFrame(columns=waiters)
        rotation.index.name = "Date"
        rotation.to_excel("./Data/Rotation.xlsx")
for waiter, waiter_history in waiter_history_dict.items():
    if waiter_history[-2] in high_capacity_sections or waiter_history[:-2] in unimportant_sections:
        priority_waiters.append(waiter)
# ------------------------------------------- FUNCTIONS ------------------------------------------- #

def RotationLog(section_assignments):
    waiter_assignments = {}
    for section_key, section_values in section_assignments.items():
        if section_values == None:
            section_values = []
        for value in section_values:
            waiter_assignments.update({f"{value}": section_key})
    new_row = pd.DataFrame(waiter_assignments, index=[today])
    rotation_df = pd.read_excel("./Data/RotationCopy.xlsx", index_col=0)
    rotation_df.index.name = "Date"
    rotation_df = pd.concat([rotation_df, new_row])
    rotation_df.to_excel("./Data/RotationCopy.xlsx", index_label="Date", sheet_name="Cycle")
    
    # -------------------- color the cells -------------------- #

    styleWorkbook = openpyxl.load_workbook("./Data/RotationCopy.xlsx")
    styleSheet = styleWorkbook["Cycle"]
    for column in styleSheet.iter_cols():
        for cell in column:
            if cell.value not in unimportant_sections:
                if cell.value in inside_sections:
                    styleSheet[cell.coordinate].fill = PatternFill(start_color="AEF359", end_color="AEF359", fill_type="solid")
                elif cell.value in outside_sections:
                    styleSheet[cell.coordinate].fill = PatternFill(start_color="BC544B", end_color="BC544B", fill_type="solid")
    styleWorkbook.save("./Data/RotationCopy.xlsx")

def CheckAvailability(waiter, section_values, section_key):
    last_sec = waiter_history_dict[waiter][-1]
    last_sec_loc = sections[last_sec]["location"]
    wanted_sec_loc = section_values["location"]
    waiter_history_list = waiter_history_dict[waiter]
    waiter_history_list.reverse()
    print(f"waiters last section is : {last_sec}. and has worked in {last_sec_loc}, looking for : {wanted_sec_loc}")
    if wanted_sec_loc == last_sec_loc:
        print("FAILED")
        return False
    print(f"capacity is : {section_values['capacity']}, waiter has worked : {waiter_history_list.count(section_key)} times.")
    if section_values["capacity"] == waiter_history_list.count(section_key):
        print("FAILED")
        return False
    print(f"waiters last {wanted_sec_loc} section is {waiter_history_list[:-2]}")
    if not section_key in waiter_history_list[:-2] and (waiter in priority_waiters or all(item in priority_waiters for item in assigned_waiters)):
        return True
    else:
        return False

# ------------------------------------------- MAIN ------------------------------------------- #
i = 0
while len(assigned_waiters) < (important_section_capacity + unimportant_section_capacity) and i < 2:
    i += 1
    for section_key, section_values in sections.items():
        if section_key in full_sections:
            continue
        for waiter in waiters:
            if waiter in assigned_waiters or section_values['capacity'] == 0:
                if section_values['capacity'] == 0:
                    full_sections.append(section_key)
                    break
                continue
            if not CheckAvailability(waiter, section_values, section_key):
                continue
            if section_values["capacity"] > 0 and section_assignments[f"{section_key}"] == None:
                section_assignments.update({f"{section_key}": [waiter]})
                assigned_waiters.append(waiter)
                section_values["capacity"] -= 1
                continue
            elif section_values["capacity"] > 0 and section_assignments[f"{section_key}"] != None:
                already_assigned = section_assignments[f"{section_key}"]
                already_assigned.append(waiter)
                section_assignments.update({f"{section_key}": already_assigned})
                assigned_waiters.append(waiter)
                section_values['capacity'] -= 1
                continue
            else:
                break

RotationLog(section_assignments)